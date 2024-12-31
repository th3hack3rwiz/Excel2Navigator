#!/usr/bin/env python3
import requests
import pandas as pd
import json
import copy
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import Counter
import matplotlib.colors as mcolors
import sys
import pyfiglet

def display_banner():
    banner = pyfiglet.figlet_format("Excel2Navigator", font="doom")
    print(banner)
    print(" " * 50 + "© th3hack3rwiz\n")

def process_single_sheet(excel_file_path):
    # Load the master JSON (layer.json)
    master_json = {}
    with open("layer.json", "r") as json_file:
        master_json = json.load(json_file)

    # Load the Excel file into a Pandas DataFrame (assuming only one sheet)
    excel_data = pd.read_excel(excel_file_path, header=None, sheet_name=None)

    # Initialize a dictionary to store entry frequencies and associated sheets
    entry_freq = {}
    ttp_sources = {}

    # Iterate over each sheet in the Excel file
    for sheet_name, sheet_data in excel_data.items():
        # Extract the first column (technique IDs) and second column (sources)
        known_techniques = sheet_data.iloc[:, 0].tolist()
        sources = sheet_data.iloc[:, 1].tolist()

        # Count the frequency of each technique
        for ttp, source in zip(known_techniques, sources):
            if ttp not in entry_freq:
                entry_freq[ttp] = 0
            entry_freq[ttp] += 1

            # Store the sources for each technique
            if ttp not in ttp_sources:
                ttp_sources[ttp] = []
            ttp_sources[ttp].append(f"{sheet_name}: {source}")

        # Filter techniques that are in the known techniques list
        copy_json = copy.deepcopy(master_json)
        filtered_techs = [t for t in copy_json["techniques"] if t["techniqueID"].lower() in known_techniques]

        # Iterate over the filtered techniques and assign scores
        for item in filtered_techs:
            technique_id = item["techniqueID"].lower()

            # Assign the score based on the frequency of the technique (1 or higher)
            if technique_id in entry_freq:
                item["score"] = 1

                # Add a comment only if the technique has a corresponding source (and a score of 1)
                if technique_id in ttp_sources:
                    item["comment"] = '\n'.join(ttp_sources[technique_id])

        # Update the techniques in the copy of the master JSON
        copy_json["techniques"] = filtered_techs

        # Save the modified JSON to a file named after the sheet
        output_path = f"{sheet_name}.json"
        with open(output_path, "w") as op:
            json.dump(copy_json, op, indent=4)

        print(f"[+] Processed sheet: {sheet_name}, ATT&CK Navigator file created and saved to {output_path}")

def nameTTPs(excel_file_path):
    # Create a dictionary to store variable-value pairs from the TTP code-name index.txt file
    variable_dict = {}

    # Read the TTP code-name index.txt file and parse variable-value pairs
    with open("TTP code-name index.txt", "r") as test_file:
        test_data = test_file.readlines()
        for line in test_data:
            parts = line.strip().split(":")
            if len(parts) == 2:
                variable = parts[0].strip().upper()
                value = parts[1].strip()
                variable_dict[variable] = value

    # Parse JSON to get map from techniqueID to tactic:
    tidToTacticMap = {}
    with open("layer.json", "r") as json_file:
        parsed_json = json.load(json_file)
        techniques = parsed_json["techniques"]

        for technique in techniques:
            tidToTacticMap[technique["techniqueID"]] = " ".join(word.capitalize() for word in technique["tactic"].replace('-', ' ').split())

    # Load the entire workbook using openpyxl
    workbook = load_workbook(excel_file_path)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Read the current sheet data
        excel_data = pd.DataFrame(workbook[sheet_name].values)
        # Optionally set the first row as the header
        excel_data.columns = excel_data.iloc[0]
        excel_data = excel_data[1:]

        # Initialize lists to store the updated values for the "Techniques" and "Tactics" columns
        updated_techniques = []
        updated_tactics = []

        # Iterate through the rows of the Excel data
        for index, row in excel_data.iterrows():
            # Get the value from "Technique ID" (Column A)
            value = row["Technique ID"]
            
            # Capitalize the value
            capitalized_value = value.upper()

            # Check if the capitalized value is in the dictionary (case-insensitive)
            if capitalized_value in variable_dict:
                updated_techniques.append(variable_dict[capitalized_value])
            else:
                updated_techniques.append("TTP Outdated In MITRE")

            # Check if this TechniqueID has a tactic
            if capitalized_value in tidToTacticMap:
                updated_tactics.append(tidToTacticMap[capitalized_value])
            else:
                updated_tactics.append("TTP Outdated In MITRE")

        # Insert new columns "Techniques" and "Tactics" with the updated values
        excel_data.insert(loc=1, column="Techniques", value=updated_techniques)
        excel_data.insert(loc=2, column="Tactics", value=updated_tactics)

        # Save the updated data back to the same sheet in the Excel file
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            excel_data.to_excel(writer, sheet_name=sheet_name, index=False)

    print("[+] Mapping TTP codes to their respective Tactics and Techniques...")

def clean_TTPs(excel_file_path):
    # Load the Excel file
    xls = pd.ExcelFile(excel_file_path)

    # Initialize an empty list to store DataFrames for each sheet
    all_dfs = []

    # Iterate over each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        # Read the sheet into a DataFrame without header
        df = pd.read_excel(xls, sheet_name, header=None)

        # Ensure there are at least two columns
        if df.shape[1] < 2:
            print(f"[-] Skipping sheet {sheet_name} as it does not have at least two columns containing TTP codes and Sources.")
            continue

        # Convert columns to string type to avoid any issues with join operation
        df[0] = df[0].astype(str)
        df[1] = df[1].astype(str)

        # Group by the first column (TTPs) and aggregate the second column (Source) as a comma-separated string
        df = df.groupby(0, as_index=False).agg(lambda x: ', '.join(x))

        # Append the DataFrame to the list
        all_dfs.append((sheet_name, df))

    # Save the result back to the same Excel file with separate sheets
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        for sheet_name, df in all_dfs:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=None)

    print(f"[+] Removing redundant TTP codes in each sheet while preserving every source... ")

def create_summary_sheet(excel_file_path):
    # Load the Excel file into a Pandas DataFrame
    excel_data = pd.read_excel(excel_file_path, header=None, sheet_name=None, engine='openpyxl')

    # Initialize a dictionary to store entry frequencies and associated sheets
    entry_freq = {}

    # Iterate over each sheet in the Excel file
    for sheet_name, sheet_data in excel_data.items():
        # Extract the first column (column index 0) as a Series
        first_column = sheet_data.iloc[:, 0]

        # Count the frequency of each entry in the first column
        freq = first_column.value_counts().to_dict()

        # Update the entry_freq dictionary with the frequencies and associated sheets from the current sheet
        for entry, frequency in freq.items():
            if entry not in entry_freq:
                entry_freq[entry] = {'frequency': frequency, 'sheets': [sheet_name]}
            else:
                entry_freq[entry]['frequency'] += frequency
                entry_freq[entry]['sheets'].append(sheet_name)

    # Sort the entries based on their frequencies in descending order
    sorted_entries = sorted(entry_freq.items(), key=lambda x: x[1]['frequency'], reverse=True)

    # Create a list to hold the data for the new sheet
    new_sheet_data = []

    # Populate the new_sheet_data list with the sorted entries
    for entry, data in sorted_entries:
        frequency = data['frequency']
        sheets = ', '.join(data['sheets'])
        new_sheet_data.append([entry, frequency, sheets])

    # Convert the new_sheet_data list into a DataFrame
    new_sheet_df = pd.DataFrame(new_sheet_data, columns=['Technique ID', 'Score', 'Actor name'])

    # Load the existing workbook
    book = load_workbook(excel_file_path)

    # Add the new sheet with the DataFrame data
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        writer._book = book
        new_sheet_df.to_excel(writer, sheet_name='Summary', index=False)

    # Reorder the sheets to make the "Summary" sheet the first one
    book._sheets.insert(0, book._sheets.pop(-1))
    book.save(excel_file_path)

    print("[+] Creating 'Summary' sheet as the first sheet in the Excel workbook.")

def add_header_row(excel_file_path):
    # Load the workbook
    book = load_workbook(excel_file_path)

    # Iterate over all sheets except the "Summary" sheet
    for sheet_name in book.sheetnames:
        if sheet_name != "Summary":
            sheet = book[sheet_name]
            # Insert a new row at the top
            sheet.insert_rows(1)
            # Set the values for cell A1 and B1
            sheet['A1'] = "Technique ID"
            sheet['B1'] = "Source"

    # Save the updated workbook
    book.save(excel_file_path)
    print("[+] Adding required headers...")

def generate_gradient(start_color, end_color, num_colors):
    cmap = mcolors.LinearSegmentedColormap.from_list('custom_gradient', [start_color, end_color])
    num_steps = num_colors
    gradient = [mcolors.to_hex(cmap(i / (num_steps - 1))) for i in range(num_steps)]
    return gradient

def clubjson(excel_file_path):
    master_json = {}
    with open("layer.json", "r") as json_file:
        master_json = json.load(json_file)
    master_json2 = copy.deepcopy(master_json)

    # Load the Excel file into a Pandas DataFrame
    excel_data = pd.read_excel(excel_file_path, header=None, sheet_name=None, engine='openpyxl')

    # Initialize a dictionary to store entry frequencies and associated sheets
    entry_freq = {}

    # Initialize a dictionary to store TTP ID to sheet sources mapping
    ttp_sources = {}

    for sheet_name, sheet_data in excel_data.items():
        known_techniques = sheet_data.iloc[:, 0].tolist()
        sources = sheet_data.iloc[:, 1].tolist()

        freq = Counter(known_techniques)

        for ttp, source in zip(known_techniques, sources):
            if ttp not in entry_freq:
                entry_freq[ttp] = 0
            entry_freq[ttp] += 1

            if ttp not in ttp_sources:
                ttp_sources[ttp] = []
            ttp_sources[ttp].append(f"{sheet_name}: {source}")

    known_techs = entry_freq.keys()

    filtered = [t for t in master_json2["techniques"] if t["techniqueID"].lower() in known_techs]

    for item in filtered:
        item["score"] = entry_freq[item["techniqueID"].lower()]
        item["comment"] = '\n'.join(ttp_sources.get(item["techniqueID"].lower(), []))

    # Determine the max frequency and generate the gradient colors
    max_value = max(entry_freq.values(), default=1)
    num_colors = max_value
    colors = generate_gradient("#8ec843", "#ff6666", num_colors)  # Light green to medium red

    master_json2["techniques"] = filtered
    master_json2["gradient"]["colors"] = colors
    master_json2["gradient"]["minValue"] = 1
    master_json2["gradient"]["maxValue"] = max_value

    with open("Heatmap.json", "w") as op:
        json.dump(master_json2, op, indent=4)

    print("[+] Creating Heatmap.json...")

def downloadRequirements():
    # URL of the GitHub Gist for layer.json file containing all TTPs from MITRE ATT&CK Navigator having a score of 1
    gist_url = "https://gist.githubusercontent.com/th3hack3rwiz/a06b5d9096fd6ca1698033a0007c5902/raw/05ff8073d6b09226cbd2a5c7036cf494ae12cea1/gistfile1.txt"

    # Make a request to fetch the content
    response = requests.get(gist_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Save the content to a file
        with open("layer.json", "w") as file:
            file.write(response.text)
        print("[+] layer.json fetched successfully.")
    else:
        print(f"[-] Failed to fetch the gist. Status code: {response.status_code}")

    # URL of the GitHub Gist for TTP code-name index.txt file containing all TTPs from MITRE ATT&CK Navigator having a score of 1
    gist_url = "https://gist.githubusercontent.com/th3hack3rwiz/1a03059bedbad5a3106ea10a1a2233b1/raw/e84e2c5c2cbdaa19518a7a22381af11d6d753924/gistfile1.txt"

    # Make a request to fetch the content
    response = requests.get(gist_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Save the content to a file
        with open("TTP code-name index.txt", "w") as file:
            file.write(response.text)
        print("[+] TTP code-name index.txt file saved successfully.")
    else:
        print(f"[-] Failed to fetch the gist. Status code: {response.status_code}")

def main():
    display_banner()
    parser = argparse.ArgumentParser(description='Excel2Navigator - Convert Excel TTPs to MITRE ATT&CK Navigator layers')
    parser.add_argument('-x', '--excel', help='Path/to/the/Excel/file/containing/TTPs', required=True)
    args = parser.parse_args()


    if not args.excel:
        parser.print_help()
        sys.exit(1)

    excel_file_path = args.excel

    # Check if the Excel file has only one sheet
    wb = load_workbook(excel_file_path)
    if len(wb.sheetnames) == 1:
        print("[+] Single sheet detected. Processing accordingly...")
        
        # Download requirements
        downloadRequirements()
        
        # Clean the TTPs
        clean_TTPs(excel_file_path)
        
        # Process single sheet and create JSON
        process_single_sheet(excel_file_path)
        
        # Add header row to the sheet
        add_header_row(excel_file_path)
        
        # Add names of Tactics and Techniques
        nameTTPs(excel_file_path)
    else:
        print("[+] Multiple sheets detected. Processing using original workflow...")
        
        # Download requirements for adding the names of Tactics and Techniques associated with TTP codes
        downloadRequirements()
        
        # Clean the TTPs
        clean_TTPs(excel_file_path)
        
        # Generate the clubbed JSON
        clubjson(excel_file_path)
        
        # Create the summary sheet
        create_summary_sheet(excel_file_path)
        
        # Add header rows to all sheets except "Summary"
        add_header_row(excel_file_path)

        # Add names of Tactics and Techniques associated with TTP codes
        nameTTPs(excel_file_path)

if __name__ == "__main__":
    main()
